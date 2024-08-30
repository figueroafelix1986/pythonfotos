from controllers.Cdate_active import DateActive, ControllersDateActive
from controllers.Cdate_work import DateWork, ControllersDateWork
from controllers.Cemployee import ControllersEmployee
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from tkinter import messagebox
from openpyxl import load_workbook
import openpyxl



class ReportesRango:
    def __init__(self):   
        self.controller_dateactive=ControllersDateActive()
        self.rango_fecha=self.controller_dateactive.listar_date_active()  
        self.controller_date_range= ControllersDateWork()
        self.controller_name=ControllersEmployee()  
        pass 

    def list_rangofecha(self):
        data = []
        for date_r in self.rango_fecha:
            fecha_ini = date_r.date_ini.strftime('%Y-%m-%d')
            fecha_fin = date_r.dete_fin.strftime('%Y-%m-%d')
            
            objeto_rango = self.controller_date_range.listar_date_work_range(fecha_ini, fecha_fin)
            for obj in objeto_rango:
                nombre_trabajador = self.controller_name.get_trabajador_id(obj.employee_id)
                data.append({
                    'nombre': nombre_trabajador,
                    'date': obj.date,
                    'work': obj.work,
                    'salaries': obj.salaries
                })

        self.export_to_excel(data, fecha_ini, fecha_fin)

    def export_to_excel(self, data, fecha_ini, fecha_fin):
        # Crear un DataFrame a partir de los datos
        df = pd.DataFrame(data)

        # Convertir la columna 'salaries' a numérica sin redondear
        df['salaries'] = pd.to_numeric(
            df['salaries'], errors='coerce').round(2)

        # Obtener el rango de fechas
        start_date = datetime.strptime(fecha_ini, '%Y-%m-%d')
        end_date = datetime.strptime(fecha_fin, '%Y-%m-%d')
        date_range = pd.date_range(start=start_date, end=end_date)

        # Crear un DataFrame con las columnas necesarias
        columns = ['nombre', 'Salario', 'a cobrar'] + [date.strftime('%d') for date in date_range]
        result_df = pd.DataFrame(columns=columns)
        result_df_salaries = pd.DataFrame(columns=columns)

        # Rellenar el DataFrame original
        for name, group in df.groupby('nombre'):
            row = {'nombre': name, 'Salario': group['salaries'].iloc[0]}
            row.update({date.strftime('%d'): '' for date in date_range})
            for _, entry in group.iterrows():
                row[entry['date'].strftime('%d')] = entry['work']
            row['a cobrar'] = group.loc[group['work'].isin(
                ['X', '2']), 'salaries'].sum()
            result_df = pd.concat([result_df, pd.DataFrame([row])], ignore_index=True)

        # Rellenar el nuevo DataFrame con los valores de salaries
        for name, group in df.groupby('nombre'):
            row_salaries = {'nombre': name,
                            'Salario': group['salaries'].iloc[0]}
            row_salaries.update(
                {date.strftime('%d'): '' for date in date_range})
            for _, entry in group.iterrows():
                row_salaries[entry['date'].strftime('%d')] = entry['salaries'] if (
                    entry['work'] == 'X' or entry['work'] == '2') else entry['work']
            row_salaries['a cobrar'] = group.loc[group['work'].isin(
                ['X', '2']), 'salaries'].sum()
            result_df_salaries = pd.concat(
                [result_df_salaries, pd.DataFrame([row_salaries])], ignore_index=True)

        # Exportar a Excel
        report_date = datetime.now().strftime('%Y%m%d')
        file_name = f"salario_{report_date}.xlsx"
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='Asistencia', index=False)
            result_df_salaries.to_excel(
                writer, sheet_name='Salaries', index=False)

        # Insertar fila con los días de la semana
        workbook = load_workbook(file_name)
        for sheet_name in ['Asistencia', 'Salaries']:
            worksheet = workbook[sheet_name]
            days_of_week = ['', '', ''] + \
                [date.strftime('%a')[:2] for date in date_range]
            worksheet.insert_rows(1)
            for col_num, value in enumerate(days_of_week, 1):
                worksheet.cell(row=1, column=col_num, value=value)

        workbook.save(file_name)
        workbook.close()

        # Ajustar el ancho de las celdas
        self.adjust_column_width(file_name, date_range, sheet_names=[
                                 'Asistencia', 'Salaries'])
        messagebox.showinfo("Reporte", f"Reporte exportado a {file_name}")


    def adjust_column_width(self, file_name, date_range, sheet_names):
        workbook = openpyxl.load_workbook(file_name)
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 5)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        workbook.save(file_name)



